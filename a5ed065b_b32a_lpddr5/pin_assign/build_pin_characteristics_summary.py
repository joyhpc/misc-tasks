#!/usr/bin/env python3
"""Build an extensible LPDDR5 pin characteristics workbook.

The workbook is a review asset, not a source of truth. Inputs remain:
- FPGA-side OrCAD pin assignment workbook.
- Memory-side package pin/net TSV files.
"""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent
FPGA_XLSX = ROOT / "fpga_side" / "fpga_pin_assign_orcad_format_hsio2b_3a_lpddr5.xlsx"
MEMORY_TSVS = {
    "U0": ("Bank2B", ROOT / "memory_side" / "lpddr5_315_DR_DS_U0_Bank2B_pin_net.tsv"),
    "U1": ("Bank3A", ROOT / "memory_side" / "lpddr5_315_DR_DS_U1_Bank3A_pin_net.tsv"),
}
OUTPUT_XLSX = ROOT / "pin_characteristics_summary.xlsx"

EXPECTED_FPGA_ONLY = {"FPGA_RZQ", "REFCLK_P", "REFCLK_N"}


@dataclass(frozen=True)
class MemoryPin:
    unit: str
    bank: str
    group: str
    memory_ball: str
    net_name: str
    memory_pin: str
    note: str


@dataclass(frozen=True)
class FpgaPin:
    unit: str
    bank: str
    side: str
    row: int
    fpga_pin: str
    pin_index: int
    signal: str | None
    net_name: str | None


def signal_class(net_name: str | None, signal: str | None = None) -> str:
    text = net_name or signal or ""
    if "ZQ" in text and "240R" in text:
        return "MEMORY_LOCAL_ZQ"
    if "FPGA_RZQ" in text:
        return "FPGA_RZQ"
    if "REFCLK" in text:
        return "REFCLK"
    if re.search(r"_(?:RDQS)\d+_", text) or re.search(r"^RDQS\d+", text):
        return "RDQS"
    if re.search(r"_(?:WCK)\d+_", text) or re.search(r"^WCK\d+", text):
        return "WCK"
    if re.search(r"_CK_[TC]$", text) or re.search(r"^CK[_ ]?", text):
        return "CK"
    if re.search(r"_(?:DQ)\d+$", text) or re.search(r"^DQ\d+", text):
        return "DQ"
    if re.search(r"_(?:DMI)\d+$", text) or re.search(r"^DMI\d+", text):
        return "DMI"
    if re.search(r"_(?:CA)\d+$", text) or re.search(r"^CA\d+", text):
        return "CA"
    if re.search(r"_(?:CS)\d+$", text) or re.search(r"^CS\d+", text):
        return "CS"
    if "RESET_N" in text:
        return "RESET_N"
    if not text:
        return "UNASSIGNED"
    return "OTHER"


def signal_index(net_name: str | None, signal: str | None = None) -> int | None:
    text = net_name or signal or ""
    for pattern in (
        r"(?:DQ|DMI|RDQS|WCK|CA|CS)(\d+)",
        r"REFCLK_([PN])",
    ):
        match = re.search(pattern, text)
        if match and match.group(1).isdigit():
            return int(match.group(1))
    return None


def diff_role(net_name: str | None, memory_pin: str | None = None) -> str:
    text = "_".join(part for part in (net_name, memory_pin) if part)
    if re.search(r"(?:_T|_t)(?:_|$)", text):
        return "T"
    if re.search(r"(?:_C|_c)(?:_|$)", text):
        return "C"
    if text.endswith("_P"):
        return "P"
    if text.endswith("_N"):
        return "N"
    return ""


def channel_from_memory_pin(memory_pin: str | None) -> str:
    if not memory_pin:
        return ""
    if memory_pin.endswith("_A"):
        return "A"
    if memory_pin.endswith("_B"):
        return "B"
    if memory_pin == "RESET_N":
        return "common"
    return ""


def byte_lane(net_name: str | None, memory_pin: str | None = None, signal: str | None = None) -> str:
    cls = signal_class(net_name, signal)
    idx = signal_index(net_name, signal)
    if cls == "DQ" and idx is not None:
        return f"byte{idx // 8}"
    if cls in {"DMI", "RDQS"} and idx is not None:
        return f"byte{idx}"
    if cls == "WCK" and idx is not None:
        return f"wck{idx}"
    if memory_pin:
        ch = channel_from_memory_pin(memory_pin)
        if ch in {"A", "B"}:
            return f"channel_{ch}"
    return ""


def route_class(cls: str) -> str:
    if cls in {"DQ", "DMI", "RDQS"}:
        return "byte_lane_data"
    if cls in {"CK", "WCK"}:
        return "differential_clock"
    if cls in {"CA", "CS", "RESET_N"}:
        return "command_address_control"
    if cls in {"FPGA_RZQ", "MEMORY_LOCAL_ZQ"}:
        return "reference_resistor"
    if cls == "REFCLK":
        return "fpga_reference_clock"
    if cls == "UNASSIGNED":
        return "unused_or_non_lpddr5"
    return "other"


def expected_memory_pin_count(cls: str) -> int:
    if cls in {"CA", "CK", "WCK", "CS"}:
        return 2
    if cls in {"DQ", "DMI", "RDQS", "RESET_N"}:
        return 1
    if cls in {"FPGA_RZQ", "REFCLK"}:
        return 0
    if cls == "MEMORY_LOCAL_ZQ":
        return 1
    return 0


def unit_from_net(net_name: str | None) -> str:
    if not net_name:
        return ""
    match = re.search(r"LP5_(U\d)_", net_name)
    return match.group(1) if match else ""


def normalize_signal_suffix(net_name: str) -> str:
    prefix = f"LP5_{unit_from_net(net_name)}_"
    return net_name.replace(prefix, "", 1)


def read_memory_pins() -> list[MemoryPin]:
    pins: list[MemoryPin] = []
    for unit, (bank, path) in MEMORY_TSVS.items():
        for line in path.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            parts = line.split("\t")
            if len(parts) < 4:
                raise ValueError(f"Bad TSV row in {path}: {line}")
            group, memory_ball, net_name, memory_pin = parts[:4]
            note = "\t".join(parts[4:]) if len(parts) > 4 else ""
            pins.append(MemoryPin(unit, bank, group, memory_ball, net_name, memory_pin, note))
    return pins


def read_fpga_pins() -> list[FpgaPin]:
    wb = load_workbook(FPGA_XLSX, read_only=True, data_only=True)
    ws = wb["check_detail"]
    pins: list[FpgaPin] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        bank, side, row_num, fpga_pin, pin_index, signal, net_name = row
        if not bank:
            continue
        unit = {"2B": "U0", "3A": "U1"}.get(str(bank), "")
        pins.append(
            FpgaPin(
                unit=unit,
                bank=str(bank),
                side=str(side),
                row=int(row_num),
                fpga_pin=str(fpga_pin),
                pin_index=int(pin_index),
                signal=str(signal) if signal else None,
                net_name=str(net_name) if net_name else None,
            )
        )
    return pins


def autosize_and_filter(ws, table_name: str | None = None) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 60))
        ws.column_dimensions[letter].width = max(max_len + 2, 10)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if table_name and ws.max_row >= 2 and ws.max_column >= 1:
        table = Table(displayName=table_name, ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)


def append_rows(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)


def build_workbook() -> None:
    memory_pins = read_memory_pins()
    fpga_pins = read_fpga_pins()

    memory_by_net: dict[str, list[MemoryPin]] = defaultdict(list)
    for pin in memory_pins:
        memory_by_net[pin.net_name].append(pin)

    fpga_by_net: dict[str, list[FpgaPin]] = defaultdict(list)
    for pin in fpga_pins:
        if pin.net_name:
            fpga_by_net[pin.net_name].append(pin)

    wb = Workbook()
    wb.remove(wb.active)

    readme = wb.create_sheet("README")
    readme_rows = [
        ["Workbook purpose", "LPDDR5 pin characteristics, net matching, and expandable review fields"],
        ["Output", OUTPUT_XLSX.name],
        ["FPGA input", FPGA_XLSX.relative_to(ROOT).as_posix()],
        ["Memory U0 input", MEMORY_TSVS["U0"][1].relative_to(ROOT).as_posix()],
        ["Memory U1 input", MEMORY_TSVS["U1"][1].relative_to(ROOT).as_posix()],
        ["Important rule", "Compare unique FPGA-connected nets, not memory-side physical pin rows"],
        ["Why memory rows are larger", "LPDDR5 CA/CK/WCK/CS shared x32 nets appear on both channel A and channel B balls"],
        ["FPGA-only expected nets", ", ".join(sorted(EXPECTED_FPGA_ONLY))],
        ["Memory-local expected net", "ZQ_A_240R_TO_VDDQ, local 240 ohm to VDDQ"],
    ]
    append_rows(readme, ["field", "value"], readme_rows)

    overview_rows: list[list[Any]] = []
    for unit, (bank, _) in MEMORY_TSVS.items():
        mem_unit = [pin for pin in memory_pins if pin.unit == unit]
        fpga_unit = [pin for pin in fpga_pins if pin.unit == unit and pin.net_name]
        mem_all = {pin.net_name for pin in mem_unit}
        mem_fpga = {pin.net_name for pin in mem_unit if signal_class(pin.net_name, pin.memory_pin) != "MEMORY_LOCAL_ZQ"}
        fpga_set = {pin.net_name for pin in fpga_unit if pin.net_name}
        fpga_only = sorted(fpga_set - mem_fpga)
        missing = sorted(mem_fpga - fpga_set)
        overview_rows.append(
            [
                unit,
                bank,
                len(mem_unit),
                len(mem_all),
                len(mem_fpga),
                len(fpga_unit),
                len(fpga_set),
                len(missing),
                ", ".join(missing),
                len(fpga_only),
                ", ".join(fpga_only),
                "PASS" if not missing and {normalize_signal_suffix(n) for n in fpga_only} == EXPECTED_FPGA_ONLY else "REVIEW",
            ]
        )
    ws = wb.create_sheet("overview")
    append_rows(
        ws,
        [
            "unit",
            "fpga_bank",
            "memory_physical_pin_rows",
            "memory_unique_all_nets",
            "memory_unique_fpga_connected_nets",
            "fpga_net_entries",
            "fpga_unique_nets",
            "missing_in_fpga_count",
            "missing_in_fpga",
            "fpga_only_count",
            "fpga_only_expected",
            "match_status",
        ],
        overview_rows,
    )

    net_rows: list[list[Any]] = []
    for net_name in sorted(set(memory_by_net) | set(fpga_by_net)):
        mems = memory_by_net.get(net_name, [])
        fpgas = fpga_by_net.get(net_name, [])
        sample_mem = mems[0] if mems else None
        sample_fpga = fpgas[0] if fpgas else None
        cls = signal_class(net_name, sample_mem.memory_pin if sample_mem else sample_fpga.signal if sample_fpga else None)
        exp_count = expected_memory_pin_count(cls)
        mem_pin_count = len(mems)
        unit = unit_from_net(net_name) or (sample_mem.unit if sample_mem else sample_fpga.unit if sample_fpga else "")
        channels = sorted({channel_from_memory_pin(pin.memory_pin) for pin in mems if channel_from_memory_pin(pin.memory_pin)})
        if cls == "MEMORY_LOCAL_ZQ":
            status = "MEMORY_LOCAL"
        elif mems and fpgas:
            status = "MATCH"
        elif mems and not fpgas:
            status = "MISSING_IN_FPGA"
        elif fpgas and not mems and normalize_signal_suffix(net_name) in EXPECTED_FPGA_ONLY:
            status = "FPGA_ONLY_EXPECTED"
        else:
            status = "REVIEW"
        count_status = "OK"
        if exp_count and mem_pin_count != exp_count:
            count_status = f"EXPECTED_{exp_count}_MEM_PINS"
        if cls in {"FPGA_RZQ", "REFCLK"} and mem_pin_count == 0:
            count_status = "OK_FPGA_ONLY"
        net_rows.append(
            [
                unit,
                net_name,
                cls,
                signal_index(net_name),
                diff_role(net_name),
                byte_lane(net_name),
                route_class(cls),
                "/".join(channels),
                "yes" if mem_pin_count > 1 else "no",
                mem_pin_count,
                ", ".join(pin.memory_ball for pin in mems),
                ", ".join(pin.memory_pin for pin in mems),
                len(fpgas),
                ", ".join(pin.fpga_pin for pin in fpgas),
                ", ".join(str(pin.pin_index) for pin in fpgas),
                ", ".join(pin.signal or "" for pin in fpgas),
                exp_count,
                status,
                count_status,
                "",
                "",
                "",
                "",
                "",
            ]
        )
    ws = wb.create_sheet("net_summary")
    append_rows(
        ws,
        [
            "unit",
            "net_name",
            "signal_class",
            "signal_index",
            "diff_role",
            "byte_lane_or_group",
            "route_class",
            "memory_channel_scope",
            "shared_on_memory_side",
            "memory_pin_count",
            "memory_balls",
            "memory_pins",
            "fpga_pin_count",
            "fpga_pins",
            "fpga_pin_indices",
            "fpga_signals",
            "expected_memory_pin_count",
            "net_match_status",
            "count_status",
            "impedance_rule",
            "length_match_group",
            "layout_status",
            "review_owner",
            "review_status",
        ],
        net_rows,
    )

    memory_rows = []
    for pin in memory_pins:
        cls = signal_class(pin.net_name, pin.memory_pin)
        memory_rows.append(
            [
                pin.unit,
                pin.bank,
                pin.group,
                pin.memory_ball,
                pin.memory_pin,
                pin.net_name,
                cls,
                signal_index(pin.net_name, pin.memory_pin),
                diff_role(pin.net_name, pin.memory_pin),
                channel_from_memory_pin(pin.memory_pin),
                byte_lane(pin.net_name, pin.memory_pin),
                route_class(cls),
                "no" if cls == "MEMORY_LOCAL_ZQ" else "yes",
                "yes" if len(memory_by_net[pin.net_name]) > 1 else "no",
                pin.note,
                "",
                "",
                "",
            ]
        )
    ws = wb.create_sheet("memory_pin_features")
    append_rows(
        ws,
        [
            "unit",
            "fpga_bank",
            "memory_group",
            "memory_ball",
            "memory_pin",
            "net_name",
            "signal_class",
            "signal_index",
            "diff_role",
            "memory_channel",
            "byte_lane_or_group",
            "route_class",
            "connects_to_fpga",
            "shared_on_memory_side",
            "source_note",
            "package_family",
            "layout_note",
            "review_status",
        ],
        memory_rows,
    )

    fpga_rows = []
    for pin in fpga_pins:
        cls = signal_class(pin.net_name, pin.signal)
        fpga_rows.append(
            [
                pin.unit,
                pin.bank,
                pin.side,
                pin.row,
                pin.fpga_pin,
                pin.pin_index,
                pin.signal or "",
                pin.net_name or "",
                cls,
                signal_index(pin.net_name, pin.signal),
                diff_role(pin.net_name, pin.signal),
                byte_lane(pin.net_name, signal=pin.signal),
                route_class(cls),
                "yes" if pin.net_name else "no",
                "yes" if pin.net_name and pin.net_name in memory_by_net else "no",
                "yes" if pin.net_name and normalize_signal_suffix(pin.net_name) in EXPECTED_FPGA_ONLY else "no",
                "",
                "",
                "",
            ]
        )
    ws = wb.create_sheet("fpga_pin_features")
    append_rows(
        ws,
        [
            "unit",
            "bank",
            "side",
            "orcad_row",
            "fpga_pin",
            "pin_index",
            "fpga_signal",
            "net_name",
            "signal_class",
            "signal_index",
            "diff_role",
            "byte_lane_or_group",
            "route_class",
            "assigned_lpddr5_net",
            "has_memory_side_pin",
            "fpga_only_expected",
            "bank_io_rule",
            "layout_note",
            "review_status",
        ],
        fpga_rows,
    )

    class_rows: list[list[Any]] = []
    for unit in sorted(MEMORY_TSVS):
        mem_unit = [pin for pin in memory_pins if pin.unit == unit]
        fpga_unit = [pin for pin in fpga_pins if pin.unit == unit and pin.net_name]
        for cls in sorted({signal_class(pin.net_name, pin.memory_pin) for pin in mem_unit} | {signal_class(pin.net_name, pin.signal) for pin in fpga_unit}):
            mem_nets = {pin.net_name for pin in mem_unit if signal_class(pin.net_name, pin.memory_pin) == cls}
            fpga_nets = {pin.net_name for pin in fpga_unit if signal_class(pin.net_name, pin.signal) == cls}
            class_rows.append(
                [
                    unit,
                    cls,
                    len([pin for pin in mem_unit if signal_class(pin.net_name, pin.memory_pin) == cls]),
                    len(mem_nets),
                    len(fpga_nets),
                    ", ".join(sorted(mem_nets - fpga_nets)),
                    ", ".join(sorted(fpga_nets - mem_nets)),
                ]
            )
    ws = wb.create_sheet("count_by_class")
    append_rows(
        ws,
        [
            "unit",
            "signal_class",
            "memory_physical_pin_rows",
            "memory_unique_nets",
            "fpga_unique_nets",
            "missing_in_fpga",
            "fpga_only",
        ],
        class_rows,
    )

    shared_rows = []
    for net_name, pins in sorted(memory_by_net.items()):
        if len(pins) <= 1:
            continue
        cls = signal_class(net_name, pins[0].memory_pin)
        shared_rows.append(
            [
                pins[0].unit,
                net_name,
                cls,
                len(pins),
                ", ".join(pin.memory_ball for pin in pins),
                ", ".join(pin.memory_pin for pin in pins),
                "A/B shared x32 net" if cls in {"CA", "CK", "WCK", "CS"} else "review",
            ]
        )
    ws = wb.create_sheet("shared_net_detail")
    append_rows(
        ws,
        ["unit", "net_name", "signal_class", "memory_pin_count", "memory_balls", "memory_pins", "reason"],
        shared_rows,
    )

    rules = wb.create_sheet("validation_rules")
    append_rows(
        rules,
        ["rule", "meaning", "action"],
        [
            [
                "memory rows != fpga nets",
                "Memory-side TSV has physical balls. Shared CA/CK/WCK/CS nets appear on channel A and B balls.",
                "Compare unique FPGA-connected net names, not TSV line count.",
            ],
            [
                "FPGA-only nets",
                "FPGA_RZQ and REFCLK_P/N are valid FPGA-side nets with no LPDDR5 component pin.",
                "Keep them in fpga_side; do not add fake memory pins.",
            ],
            [
                "Memory-local ZQ",
                "LPDDR5 ZQ_A uses local 240 ohm to VDDQ and does not connect to FPGA_RZQ.",
                "Keep it on memory-side only.",
            ],
            [
                "Extensible review columns",
                "net_summary includes blank columns for impedance, length group, layout and review ownership.",
                "Fill those during schematic/layout review; rerun script if source pin tables change.",
            ],
        ],
    )

    for idx, ws in enumerate(wb.worksheets, start=1):
        table_name = re.sub(r"[^A-Za-z0-9_]", "_", ws.title)
        if table_name[0].isdigit():
            table_name = f"T_{table_name}"
        autosize_and_filter(ws, f"tbl_{idx}_{table_name}")
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(OUTPUT_XLSX)
    print(f"Wrote {OUTPUT_XLSX.relative_to(ROOT.parent.parent)}")


if __name__ == "__main__":
    build_workbook()
