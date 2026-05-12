#!/usr/bin/env python3
"""Generate an external-safe DDR memory supplier inquiry workbook."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUT_PATH = Path(__file__).with_name("ddr_memory_supplier_inquiry_template.xlsx")

HEADERS = [
    "方案类型",
    "推荐优先级",
    "推荐料号",
    "品牌",
    "容量",
    "位宽/Package Width",
    "Die Organization",
    "每板建议颗数",
    "每板总位宽/容量",
    "速率等级",
    "封装",
    "温度等级",
    "生命周期状态",
    "替代料号",
    "替代料兼容性",
    "样品/量产交期",
    "MOQ/MPQ",
    "样品价",
    "1k价",
    "5k价",
    "资料是否齐全",
    "备注/信息冲突说明",
]

INQUIRY_ROWS = [
    [
        "DDR4",
        "最低成本首选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x16 或 x8，组成 64bit 总位宽",
        "不适用/待确认",
        "待确认",
        "64bit / 容量待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "如有 NRND/EOL/LTB 风险必须提供",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "优先当前成本最低方案；如 DDR4 当前价格高于 DDR5，请明确说明",
    ],
    [
        "DDR4",
        "长期供货备选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x16 或 x8，组成 64bit 总位宽",
        "不适用/待确认",
        "待确认",
        "64bit / 容量待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "如有 NRND/EOL/LTB 风险必须提供",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "优先生命周期稳定、可持续供货、有明确替代料的方案",
    ],
    [
        "DDR4",
        "供应链替代备选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x16 或 x8，组成 64bit 总位宽",
        "不适用/待确认",
        "待确认",
        "64bit / 容量待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "必须填写",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "作为最低成本或长期供货方案的替代路线，请说明替代限制",
    ],
    [
        "DDR5",
        "成本/带宽折中",
        "待确认",
        "待确认",
        "待确认",
        "优先 x16 或 x8，组成 64bit 总位宽",
        "不适用/待确认",
        "待确认",
        "64bit / 容量待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "如有 NRND/EOL/LTB 风险必须提供",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "请说明相对 DDR4 的成本、带宽、交期和生命周期差异",
    ],
    [
        "DDR5",
        "长期供货备选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x16 或 x8，组成 64bit 总位宽",
        "不适用/待确认",
        "待确认",
        "64bit / 容量待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "如有 NRND/EOL/LTB 风险必须提供",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "优先生命周期长、供货稳定、资料完整的方案",
    ],
    [
        "DDR5",
        "供应链替代备选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x16 或 x8，组成 64bit 总位宽",
        "不适用/待确认",
        "待确认",
        "64bit / 容量待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "必须填写",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "作为成本/带宽或长期供货方案的替代路线，请说明替代限制",
    ],
    [
        "LPDDR4/LPDDR4X",
        "低功耗/面积备选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x32 package width，组成 64bit 总位宽",
        "必须填写，待确认",
        "待确认",
        "64bit / 容量待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "如有 NRND/EOL/LTB 风险必须提供",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "如果 LPDDR4 并不比 LPDDR5 便宜，请明确说明；必须说明 package width 和 die organization",
    ],
    [
        "LPDDR4/LPDDR4X",
        "长期供货备选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x32 package width，组成 64bit 总位宽",
        "必须填写，待确认",
        "待确认",
        "64bit / 容量待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "如有 NRND/EOL/LTB 风险必须提供",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "请说明是否有持续供货风险、替代料限制和资料缺口",
    ],
    [
        "LPDDR4/LPDDR4X",
        "供应链替代备选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x32 package width，组成 64bit 总位宽",
        "必须填写，待确认",
        "待确认",
        "64bit / 容量待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "必须填写",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "请说明 package width、die organization、ball map 差异是否影响替代",
    ],
    [
        "LPDDR5/LPDDR5X",
        "性能/长期供货备选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x32 package width，组成 64bit 总位宽",
        "必须填写，待确认",
        "待确认",
        "64bit / 容量待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "如有 NRND/EOL/LTB 风险必须提供",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "请说明相对 LPDDR4 的成本、性能、交期和生命周期差异；必须说明 package width 和 die organization",
    ],
    [
        "LPDDR5/LPDDR5X",
        "成本优化备选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x32 package width，组成 64bit 总位宽",
        "必须填写，待确认",
        "待确认",
        "64bit / 容量待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "如有 NRND/EOL/LTB 风险必须提供",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "如果 LPDDR5 成本已接近或低于 LPDDR4，请明确说明",
    ],
    [
        "LPDDR5/LPDDR5X",
        "供应链替代备选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x32 package width，组成 64bit 总位宽",
        "必须填写，待确认",
        "待确认",
        "64bit / 容量待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "必须填写",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "请说明 package width、die organization、ball map 差异是否影响替代",
    ],
]

SUMMARY_ROWS = [
    ["综合成本最低方案", "待填写", "需比较样品价、1k价、5k价、MOQ/MPQ、每板颗数和每板总容量"],
    ["当前供货最稳方案", "待填写", "需说明库存、样品交期、量产交期、可持续供货依据"],
    ["生命周期风险最低方案", "待填写", "需说明生命周期状态、NRND/EOL/LTB 风险、替代料路径"],
    ["最低成本方案与长期供货方案是否一致", "待填写", "如不一致，请说明价格、交期、生命周期、资料完整性或替代料风险原因"],
    ["DDR4 当前价格是否高于 DDR5", "待填写", "如高于，请明确说明报价依据、容量/位宽可比性和交期差异"],
    ["LPDDR4 当前价格是否低于 LPDDR5", "待填写", "如 LPDDR4 并不更便宜，请明确说明报价依据、容量/速率可比性和生命周期差异"],
    ["是否存在资料冲突", "待填写", "请列出 datasheet、官网、代理系统、原厂反馈之间的冲突"],
    ["是否存在不建议作为唯一推荐的方案", "待填写", "存在 NRND/EOL/LTB、资料缺失、替代路径不清晰或交期不可控时必须说明"],
]

RULE_ROWS = [
    ["保密口径", "目标 FPGA / SoC 平台，具体厂商、平台系列和型号暂不披露。"],
    ["评估边界", "本次只做存储器器件侧的成本、供货、生命周期、资料完整性和替代风险评估。"],
    ["兼容性边界", "最终主控兼容性由我方内部确认；供应商只需列出器件侧资料和风险点。"],
    ["DDR4 / DDR5 位宽", "优先推荐可组成 64bit 总位宽的 x16 或 x8 方案。"],
    ["LPDDR4 / LPDDR5 位宽", "优先推荐可组成 64bit 总位宽的 x32 package width 方案。"],
    ["LPDDR 组织信息", "LPDDR 必须明确 package width 和 die organization，不要只写 x32。"],
    ["温度等级", "商业级 / 工业级均可，但必须明确温度等级。"],
    ["生命周期", "如料号存在 NRND / EOL / LTB 风险，必须提供替代料号。"],
    ["信息缺口", "如果某项信息无法确认，不要留空，填写“待确认”或“不确定”。"],
    ["信息冲突", "如 datasheet、官网、代理系统、原厂反馈存在冲突，必须在备注中说明。"],
]

DOC_ROWS = [
    ["器件 datasheet", "必须提供", "确认容量、速率、位宽、温度等级、电源、时序和订购信息"],
    ["package drawing", "LPDDR 必须提供；DDR 建议提供", "用于确认封装尺寸、球/引脚定义和 layout 风险"],
    ["ball map / pin map", "LPDDR 必须提供；DDR 建议提供", "用于确认 package width、die organization 和替代风险"],
    ["生命周期/PCN/EOL 依据", "必须提供", "官网、原厂 PCN 或原厂代理系统优先"],
    ["代理系统截图或链接", "建议提供", "用于交期、价格、库存、MOQ/MPQ、状态交叉验证"],
    ["报价单", "必须提供", "样品价、1k价、5k价分开；注明币种、有效期、税费/运费边界"],
    ["替代料资料", "如有风险必须提供", "替代料也需提供生命周期、交期、价格和资料完整性"],
    ["质量/合规资料", "建议提供", "如 RoHS、REACH、MSL、原厂质量文件可提供，请列出"],
]

CONFLICT_ROWS = [
    ["料号状态冲突", "以原厂官网 / 原厂 PCN / 原厂代理系统为优先依据", "备注中列出冲突来源"],
    ["封装或 ball map 冲突", "必须提供 package drawing 和 ball map", "没有关键资料的方案不得作为唯一推荐"],
    ["价格冲突", "必须按样品、1k、5k 分开", "不接受单一口头价格"],
    ["生命周期冲突", "存在 NRND / EOL / LTB 风险时不能作为唯一推荐", "必须提供替代料号"],
    ["兼容性冲突", "由我方内部确认", "供应商只需列出器件侧资料和风险点"],
]


def style_sheet(ws):
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            cell.font = Font(name="Arial", size=10)


def style_header(ws, row=1, fill="1F4E78"):
    for cell in ws[row]:
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="Arial", size=10, color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_title_sheet(wb):
    ws = wb.active
    ws.title = "使用说明"
    rows = [
        ["项目", "内容"],
        ["文件用途", "向代理/供应商同时询价 DDR4 / DDR5 / LPDDR4 / LPDDR5 / LPDDR4X / LPDDR5X 存储器方案。"],
        ["比较目标", "成本、供货、生命周期、资料完整性、替代料风险、方案风险。"],
        ["对外保密口径", "目标 FPGA / SoC 平台，具体厂商、平台系列和型号暂不披露。"],
        ["评估边界", "本次只做存储器器件侧评估；最终主控兼容性由我方内部确认。"],
        ["填写要求", "所有无法确认的信息请填写“待确认”或“不确定”，不要留空。"],
        ["主表位置", "请供应商主要填写“供应商填写表”sheet。"],
        ["结论位置", "请供应商填写“结论汇总”sheet，便于直接比较四类方案。"],
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    style_header(ws)
    set_widths(ws, [22, 120])
    ws.freeze_panes = "A2"


def add_email_sheet(wb):
    ws = wb.create_sheet("对外发送正文")
    ws.append(["用途", "可直接复制发送给代理/供应商"])
    ws.append(
        [
            "正文",
            (
                "您好，\n\n"
                "请协助同时评估 DDR4 / DDR5 / LPDDR4 / LPDDR5 四类存储方案，并按本 Excel 表格反馈可推荐的器件方案、价格、交期、生命周期状态、替代料和资料完整性。\n\n"
                "主控为目标 FPGA / SoC 平台，具体厂商、平台系列和型号暂不披露。本次只做存储器器件侧的成本、供货、生命周期和资料完整性评估，最终主控兼容性由我方内部确认。\n\n"
                "请优先推荐当前可供货、生命周期风险低、资料完整且有明确替代路径的方案。如存在信息冲突或无法确认的项目，请在表格中明确标注“待确认”或“不确定”，不要留空。"
            ),
        ]
    )
    style_sheet(ws)
    style_header(ws)
    set_widths(ws, [18, 120])


def add_inquiry_sheet(wb):
    ws = wb.create_sheet("供应商填写表")
    ws.append(HEADERS)
    for row in INQUIRY_ROWS:
        ws.append(row)
    for _ in range(8):
        ws.append(["待新增"] + ["待确认"] * (len(HEADERS) - 1))
    style_sheet(ws)
    style_header(ws)
    widths = [18, 20, 24, 16, 16, 34, 28, 16, 24, 16, 18, 16, 18, 24, 28, 20, 16, 14, 14, 14, 16, 52]
    set_widths(ws, widths)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    validations = {
        "A": '"DDR4,DDR5,LPDDR4/LPDDR4X,LPDDR5/LPDDR5X,待新增"',
        "L": '"商业级,工业级,车规级,待确认,不确定"',
        "M": '"Active,NRND,EOL,LTB,待确认,不确定"',
        "U": '"齐全,部分齐全,缺失,待确认,不确定"',
    }
    for column, formula in validations.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(f"{column}2:{column}200")

    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    for col in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col)
        ws.conditional_formatting.add(
            f"{letter}2:{letter}200",
            FormulaRule(formula=[f'OR({letter}2="待确认",{letter}2="不确定")'], fill=yellow_fill),
        )


def add_rules_sheet(wb):
    ws = wb.create_sheet("评估要求")
    ws.append(["要求项", "要求内容"])
    for row in RULE_ROWS:
        ws.append(row)
    style_sheet(ws)
    style_header(ws)
    set_widths(ws, [28, 110])
    ws.freeze_panes = "A2"


def add_docs_sheet(wb):
    ws = wb.create_sheet("资料清单")
    ws.append(["资料项", "要求级别", "用途/检查点"])
    for row in DOC_ROWS:
        ws.append(row)
    style_sheet(ws)
    style_header(ws)
    set_widths(ws, [30, 26, 90])
    ws.freeze_panes = "A2"


def add_summary_sheet(wb):
    ws = wb.create_sheet("结论汇总")
    ws.append(["结论问题", "供应商结论", "必须说明的依据"])
    for row in SUMMARY_ROWS:
        ws.append(row)
    style_sheet(ws)
    style_header(ws)
    set_widths(ws, [42, 32, 90])
    ws.freeze_panes = "A2"


def add_conflict_sheet(wb):
    ws = wb.create_sheet("冲突处理规则")
    ws.append(["冲突类型", "处理规则", "备注要求"])
    for row in CONFLICT_ROWS:
        ws.append(row)
    style_sheet(ws)
    style_header(ws)
    set_widths(ws, [30, 70, 70])
    ws.freeze_panes = "A2"


def add_check_sheet(wb):
    ws = wb.create_sheet("自检")
    rows = [
        ["检查项", "结果", "说明"],
        ["没有出现具体主控厂商", "PASS", "仅使用统一保密口径"],
        ["没有出现具体平台系列", "PASS", "未披露平台系列"],
        ["没有出现具体 FPGA / SoC 型号", "PASS", "未披露芯片型号"],
        ["没有出现具体封装型号", "PASS", "封装字段留给供应商填写"],
        ["覆盖 DDR4 / DDR5 / LPDDR4 / LPDDR5", "PASS", "主表包含四类方案及扩展类型"],
        ["包含生命周期和替代料字段", "PASS", "主表包含生命周期状态、替代料号、替代料兼容性"],
        ["包含信息冲突说明字段", "PASS", "主表包含备注/信息冲突说明，另有冲突处理规则 sheet"],
        ["供应商可直接填写", "PASS", "主表有筛选、冻结表头、下拉选项和待确认高亮"],
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    style_header(ws)
    set_widths(ws, [42, 16, 90])
    ws.freeze_panes = "A2"


def main():
    wb = Workbook()
    add_title_sheet(wb)
    add_email_sheet(wb)
    add_inquiry_sheet(wb)
    add_rules_sheet(wb)
    add_docs_sheet(wb)
    add_summary_sheet(wb)
    add_conflict_sheet(wb)
    add_check_sheet(wb)
    wb.save(OUT_PATH)
    print(OUT_PATH)


if __name__ == "__main__":
    main()
