#!/usr/bin/env python3
"""Generate a supplier-facing DDR memory inquiry workbook."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUT_PATH = Path(__file__).with_name("ddr_memory_supplier_inquiry_supplier.xlsx")

MAIN_HEADERS = [
    "方案类型",
    "推荐优先级",
    "推荐料号",
    "品牌/原厂",
    "容量",
    "器件位宽/Package Width",
    "Die Organization",
    "每板建议颗数",
    "每板总位宽",
    "每板总容量",
    "速率等级",
    "工作电压/电源要求",
    "封装",
    "封装尺寸/Package Drawing",
    "温度等级",
    "生命周期状态",
    "PCN/EOL/LTB 依据",
    "替代料号",
    "替代料兼容性",
    "替代风险说明",
    "库存状态",
    "样品交期",
    "量产交期",
    "MOQ",
    "MPQ",
    "样品价",
    "1k价",
    "5k价",
    "报价币种/有效期",
    "资料完整性",
    "需提供资料",
    "备注/信息冲突说明",
]

MAIN_ROWS = [
    [
        "DDR4",
        "最低成本首选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x16 或 x8",
        "不适用/待确认",
        "待确认",
        "64bit",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "请提供 package drawing",
        "待确认",
        "待确认",
        "待确认",
        "如有 NRND/EOL/LTB 风险必须提供",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "datasheet、package drawing、生命周期/状态依据、报价单",
        "优先当前成本最低方案；如 DDR4 当前价格高于 DDR5，请明确说明",
    ],
    [
        "DDR4",
        "长期供货备选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x16 或 x8",
        "不适用/待确认",
        "待确认",
        "64bit",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "请提供 package drawing",
        "待确认",
        "待确认",
        "待确认",
        "如有 NRND/EOL/LTB 风险必须提供",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "datasheet、package drawing、生命周期/状态依据、报价单",
        "优先生命周期稳定、可持续供货、有明确替代料的方案",
    ],
    [
        "DDR4",
        "供应链替代备选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x16 或 x8",
        "不适用/待确认",
        "待确认",
        "64bit",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "请提供 package drawing",
        "待确认",
        "待确认",
        "待确认",
        "必须填写",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "datasheet、package drawing、替代料资料、报价单",
        "请说明与主推荐 DDR4 方案的替代限制",
    ],
    [
        "DDR5",
        "成本/带宽折中",
        "待确认",
        "待确认",
        "待确认",
        "优先 x16 或 x8",
        "不适用/待确认",
        "待确认",
        "64bit",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "请提供 package drawing",
        "待确认",
        "待确认",
        "待确认",
        "如有 NRND/EOL/LTB 风险必须提供",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "datasheet、package drawing、生命周期/状态依据、报价单",
        "请说明相对 DDR4 的成本、带宽、交期和生命周期差异",
    ],
    [
        "DDR5",
        "长期供货备选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x16 或 x8",
        "不适用/待确认",
        "待确认",
        "64bit",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "请提供 package drawing",
        "待确认",
        "待确认",
        "待确认",
        "如有 NRND/EOL/LTB 风险必须提供",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "datasheet、package drawing、生命周期/状态依据、报价单",
        "优先生命周期长、供货稳定、资料完整的方案",
    ],
    [
        "DDR5",
        "供应链替代备选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x16 或 x8",
        "不适用/待确认",
        "待确认",
        "64bit",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "请提供 package drawing",
        "待确认",
        "待确认",
        "待确认",
        "必须填写",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "datasheet、package drawing、替代料资料、报价单",
        "请说明与主推荐 DDR5 方案的替代限制",
    ],
    [
        "LPDDR4/LPDDR4X",
        "低功耗/面积备选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x32 package width",
        "必须填写，不要只写 x32",
        "待确认",
        "64bit",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "必须提供 package drawing 和 ball map",
        "待确认",
        "待确认",
        "待确认",
        "如有 NRND/EOL/LTB 风险必须提供",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "datasheet、package drawing、ball map、生命周期/状态依据、报价单",
        "如果 LPDDR4 并不比 LPDDR5 便宜，请明确说明",
    ],
    [
        "LPDDR4/LPDDR4X",
        "长期供货备选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x32 package width",
        "必须填写，不要只写 x32",
        "待确认",
        "64bit",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "必须提供 package drawing 和 ball map",
        "待确认",
        "待确认",
        "待确认",
        "如有 NRND/EOL/LTB 风险必须提供",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "datasheet、package drawing、ball map、生命周期/状态依据、报价单",
        "请说明是否有持续供货风险、替代料限制和资料缺口",
    ],
    [
        "LPDDR4/LPDDR4X",
        "供应链替代备选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x32 package width",
        "必须填写，不要只写 x32",
        "待确认",
        "64bit",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "必须提供 package drawing 和 ball map",
        "待确认",
        "待确认",
        "待确认",
        "必须填写",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "datasheet、package drawing、ball map、替代料资料、报价单",
        "请说明 package width、die organization、ball map 差异是否影响替代",
    ],
    [
        "LPDDR5/LPDDR5X",
        "性能/长期供货备选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x32 package width",
        "必须填写，不要只写 x32",
        "待确认",
        "64bit",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "必须提供 package drawing 和 ball map",
        "待确认",
        "待确认",
        "待确认",
        "如有 NRND/EOL/LTB 风险必须提供",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "datasheet、package drawing、ball map、生命周期/状态依据、报价单",
        "请说明相对 LPDDR4 的成本、性能、交期和生命周期差异",
    ],
    [
        "LPDDR5/LPDDR5X",
        "成本优化备选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x32 package width",
        "必须填写，不要只写 x32",
        "待确认",
        "64bit",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "必须提供 package drawing 和 ball map",
        "待确认",
        "待确认",
        "待确认",
        "如有 NRND/EOL/LTB 风险必须提供",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "datasheet、package drawing、ball map、生命周期/状态依据、报价单",
        "如果 LPDDR5 成本已接近或低于 LPDDR4，请明确说明",
    ],
    [
        "LPDDR5/LPDDR5X",
        "供应链替代备选",
        "待确认",
        "待确认",
        "待确认",
        "优先 x32 package width",
        "必须填写，不要只写 x32",
        "待确认",
        "64bit",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "必须提供 package drawing 和 ball map",
        "待确认",
        "待确认",
        "待确认",
        "必须填写",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "待确认",
        "datasheet、package drawing、ball map、替代料资料、报价单",
        "请说明 package width、die organization、ball map 差异是否影响替代",
    ],
]

SUMMARY_ROWS = [
    ["四类方案中哪一种综合成本最低？", "待填写", "请比较样品价、1k价、5k价、MOQ/MPQ、每板建议颗数和每板总容量。"],
    ["哪一种当前供货最稳？", "待填写", "请说明库存、样品交期、量产交期和代理系统/原厂反馈依据。"],
    ["哪一种生命周期风险最低？", "待填写", "请说明生命周期状态、NRND/EOL/LTB 风险、替代料路径和依据。"],
    ["最低成本方案和长期供货方案是否为同一种？", "待填写", "如不是，请说明价格、交期、生命周期、资料完整性或替代料风险原因。"],
    ["DDR4 当前价格是否高于 DDR5？", "待填写", "如高于，请明确说明报价依据、容量/位宽可比性和交期差异。"],
    ["LPDDR4 当前价格是否低于 LPDDR5？", "待填写", "如 LPDDR4 并不更便宜，请明确说明报价依据、容量/速率可比性和生命周期差异。"],
    ["是否有方案不建议作为唯一推荐？", "待填写", "存在 NRND/EOL/LTB、资料缺失、替代路径不清晰或交期不可控时必须说明。"],
    ["最终推荐排序", "待填写", "请按综合建议排序，并说明每类方案的推荐/不推荐原因。"],
]

DOC_ROWS = [
    ["器件 datasheet", "必须提供", "请提供文件名/链接", "确认容量、速率、位宽、温度等级、电源、时序和订购信息。"],
    ["package drawing", "必须提供", "请提供文件名/链接", "用于确认封装尺寸、球/引脚定义和 layout 风险。"],
    ["ball map / pin map", "LPDDR 必须提供，DDR 建议提供", "请提供文件名/链接", "用于确认 package width、die organization 和替代风险。"],
    ["生命周期/PCN/EOL 依据", "必须提供", "请提供文件名/链接", "官网、原厂 PCN 或原厂代理系统优先。"],
    ["代理系统截图或链接", "建议提供", "请提供文件名/链接", "用于交期、价格、库存、MOQ/MPQ、状态交叉验证。"],
    ["正式报价单", "必须提供", "请提供文件名/链接", "样品价、1k价、5k价分开；注明币种、有效期、税费/运费边界。"],
    ["替代料资料", "如有风险必须提供", "请提供文件名/链接", "替代料也需提供生命周期、交期、价格和资料完整性。"],
    ["质量/合规资料", "建议提供", "请提供文件名/链接", "如 RoHS、REACH、MSL、原厂质量文件可提供，请列出。"],
]

CONFLICT_ROWS = [
    ["料号状态冲突", "以原厂官网 / 原厂 PCN / 原厂代理系统为优先依据", "请在备注中列出冲突来源和推荐采用依据。"],
    ["封装或 ball map 冲突", "必须提供 package drawing 和 ball map", "没有关键资料的方案不得作为唯一推荐。"],
    ["价格冲突", "必须按样品、1k、5k 分开填写", "不接受单一口头价格；请说明报价币种、有效期和是否含税/运费。"],
    ["生命周期冲突", "若存在 NRND / EOL / LTB 风险，不能作为唯一推荐", "必须提供替代料号，并说明替代料生命周期、交期、价格和资料状态。"],
    ["兼容性冲突", "最终主控兼容性由我方确认", "供应商只需列出器件侧资料和风险点。"],
]


def style_sheet(ws):
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            cell.font = Font(name="Arial", size=10)


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(name="Arial", size=10, color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_intro_sheet(wb):
    ws = wb.active
    ws.title = "填写说明"
    rows = [
        ["项目", "供应商填写要求"],
        ["询价范围", "请同时评估 DDR4 / DDR5 / LPDDR4 / LPDDR5 四类存储方案，可包含 LPDDR4X / LPDDR5X。"],
        ["主控说明", "主控为目标 FPGA / SoC 平台，具体厂商、平台系列和型号暂不披露。"],
        ["评估边界", "本次只做存储器器件侧的成本、供货、生命周期、资料完整性和替代料风险评估。"],
        ["兼容性说明", "最终主控兼容性由我方确认；请供应商列出器件侧资料和风险点。"],
        ["DDR4 / DDR5 要求", "优先推荐可组成 64bit 总位宽的 x16 或 x8 方案。"],
        ["LPDDR4 / LPDDR5 要求", "优先推荐可组成 64bit 总位宽的 x32 package width 方案。"],
        ["LPDDR 必填项", "必须明确 package width 和 die organization，不要只写 x32。"],
        ["温度等级", "商业级 / 工业级均可，但需要明确温度等级。"],
        ["生命周期风险", "如料号存在 NRND / EOL / LTB 风险，必须提供替代料号。"],
        ["资料冲突", "如 datasheet、官网、代理系统、原厂反馈存在冲突，必须在备注中说明。"],
        ["空缺信息", "如果某项信息无法确认，不要留空，填写“待确认”或“不确定”。"],
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    style_header(ws)
    set_widths(ws, [26, 120])
    ws.freeze_panes = "A2"


def add_main_sheet(wb):
    ws = wb.create_sheet("方案报价表")
    ws.append(MAIN_HEADERS)
    for row in MAIN_ROWS:
        ws.append(row)
    for _ in range(8):
        ws.append(["待新增"] + ["待确认"] * (len(MAIN_HEADERS) - 1))

    style_sheet(ws)
    style_header(ws)
    set_widths(
        ws,
        [
            18, 20, 24, 16, 16, 26, 28, 15, 15, 16, 16, 22, 18, 28, 16, 18,
            24, 24, 28, 28, 18, 16, 16, 12, 12, 14, 14, 14, 22, 18, 38, 56,
        ],
    )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column, formula in {
        "A": '"DDR4,DDR5,LPDDR4/LPDDR4X,LPDDR5/LPDDR5X,待新增"',
        "O": '"商业级,工业级,车规级,待确认,不确定"',
        "P": '"Active,NRND,EOL,LTB,待确认,不确定"',
        "AD": '"齐全,部分齐全,缺失,待确认,不确定"',
    }.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(f"{column}2:{column}300")

    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    for col in range(1, len(MAIN_HEADERS) + 1):
        letter = get_column_letter(col)
        ws.conditional_formatting.add(
            f"{letter}2:{letter}300",
            FormulaRule(formula=[f'OR({letter}2="待确认",{letter}2="不确定")'], fill=yellow_fill),
        )


def add_summary_sheet(wb):
    ws = wb.create_sheet("结论汇总")
    ws.append(["结论问题", "供应商结论", "必须说明的依据"])
    for row in SUMMARY_ROWS:
        ws.append(row)
    style_sheet(ws)
    style_header(ws)
    set_widths(ws, [46, 32, 100])
    ws.freeze_panes = "A2"


def add_docs_sheet(wb):
    ws = wb.create_sheet("资料清单")
    ws.append(["资料项", "要求级别", "供应商填写：文件名/链接", "用途/检查点"])
    for row in DOC_ROWS:
        ws.append(row)
    style_sheet(ws)
    style_header(ws)
    set_widths(ws, [30, 26, 42, 90])
    ws.freeze_panes = "A2"


def add_conflict_sheet(wb):
    ws = wb.create_sheet("冲突处理规则")
    ws.append(["冲突类型", "处理规则", "供应商备注要求"])
    for row in CONFLICT_ROWS:
        ws.append(row)
    style_sheet(ws)
    style_header(ws)
    set_widths(ws, [30, 76, 90])
    ws.freeze_panes = "A2"


def main():
    wb = Workbook()
    add_intro_sheet(wb)
    add_main_sheet(wb)
    add_summary_sheet(wb)
    add_docs_sheet(wb)
    add_conflict_sheet(wb)
    wb.save(OUT_PATH)
    print(OUT_PATH)


if __name__ == "__main__":
    main()
